"""
camera_tab.py  [v8 – KHO HÀNG + XUẤT EXCEL]
=================================
Nâng cấp v7:
  1. Đổi tên tab Camera -> tab KHO HÀNG.
  2. Đọc nhiều QR trong nhiều khung ROI cùng 1 lúc.
  3. Cột DANH SÁCH KHO HÀNG – hiển thị QR đang có mặt trong ROI (màu xanh).
  4. Nếu QR biến mất khỏi ROI > 15 giây -> chuyển sang cột HÀNG ĐÃ LẤY (màu vàng).
Nâng cấp v8:
  5. Nút XUẤT EXCEL trên từng cột -> xuất Kho Hàng / Hàng Đã Lấy ra file .xlsx.
  6. Nút XUẤT TỔNG HỢP -> gộp cả 2 cột vào 1 file Excel 2 sheet.
"""

import re
import time
import customtkinter as ctk
from vision.cv2_safe import get_cv2
import threading
import queue
from datetime import datetime
from tkinter import filedialog, messagebox
from PIL import Image
from vision.camera_manager import CameraManager
from vision.roi_manager import ROIManager
from constants import *

_DISPLAY_W = 760
_DISPLAY_H  = 600
_FRAME_INTERVAL_MS = 33   # ~30 fps

DEFAULT_CM_SIZE   = 4.0
DEFAULT_PX_PER_CM = 30

# Thời gian chờ trước khi chuyển sang "đã lấy" (giây)
QR_ABSENT_TIMEOUT = 15


class CameraTab:

    def __init__(self, parent, shared):
        self.parent = parent
        self.shared = shared

        self.camera      = CameraManager()
        self.roi_manager = ROIManager()
        self.shared["roi_manager"] = self.roi_manager

        self.running    = False
        self.adding_roi = False
        self._last_photo = None     # giữ reference CTkImage, tránh GC gây lỗi reconnect
        self._ui_generation = 0     # tăng mỗi lần open/close để hủy vòng after cũ
        self._ui_after_id   = None  # ID của after callback đang chờ, để cancel được
        self._last_tk_photo = None  # FIX pyimage v2: hard ref tới PhotoImage thô bên trong CTkImage

        self._frame_queue: queue.Queue = queue.Queue(maxsize=1)
        self._worker_thread = None
        self._auto_mapped_qr: set = set()

        self._lock = threading.Lock()

        # ── Kho hàng (QR đang hiện diện trong ROI) ──────────────
        # qr -> {"roi", "time_first", "time_last", "count"}
        self._inventory: dict = {}
        self._inventory_order: list = []

        # ── Hàng đã lấy (QR biến mất > 15s) ────────────────────
        # qr -> {"roi", "time_in", "time_out"}
        self._taken: dict = {}
        self._taken_order: list = []

        # Theo dõi QR nào đang "vắng mặt" và khi nào bắt đầu vắng
        # qr -> timestamp lần cuối thấy
        self._last_seen: dict = {}

        # ── Drag-to-draw ROI ─────────────────────────────────────
        self._roi_drag_start: tuple | None = None   # (x, y) khi nhấn chuột
        self._roi_drag_end:   tuple | None = None   # (x, y) khi kéo
        self._roi_drawing:    bool         = False  # đang kéo vẽ
        self._last_frame_shape: tuple      = (_DISPLAY_H, _DISPLAY_W)  # (h, w) frame gốc

        self._build_ui()

    # ============================================================
    # UI
    # ============================================================

    def _build_ui(self):
        root = ctk.CTkFrame(self.parent, fg_color=C_BG)
        root.pack(fill="both", expand=True)

        # ── Top bar ──────────────────────────────────────────────
        bar = ctk.CTkFrame(root, fg_color=C_PANEL, corner_radius=10, height=56)
        bar.pack(fill="x", padx=8, pady=(8, 4))
        bar.pack_propagate(False)

        ctk.CTkLabel(bar, text="CAM KHO", font=("Consolas", 11, "bold"),
                     text_color=C_SUBTEXT).pack(side="left", padx=(14, 2))
        self.cam_menu = ctk.CTkComboBox(bar, values=[str(i) for i in range(8)],
                                        font=("Consolas", 11), width=64)
        self.cam_menu.set("0")
        self.cam_menu.pack(side="left", padx=(0, 8))

        ctk.CTkButton(bar, text="> OPEN", fg_color="#15803d", hover_color="#166534",
                      font=("Consolas", 11, "bold"), width=88, height=34,
                      command=self._open_camera).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="STOP CLOSE", fg_color=C_RED, hover_color="#991b1b",
                      font=("Consolas", 11, "bold"), width=88, height=34,
                      command=self.close_camera).pack(side="left", padx=4)

        ctk.CTkFrame(bar, width=1, fg_color=C_BORDER).pack(side="left", fill="y", padx=8, pady=8)

        ctk.CTkButton(bar, text="EDIT VẼ ROI", fg_color=C_ACCENT2, hover_color="#5b21b6",
                      font=("Consolas", 11, "bold"), width=100, height=34,
                      command=self._enable_roi).pack(side="left", padx=4)

        self.roi_menu = ctk.CTkComboBox(bar, values=["NO ROI"],
                                        font=("Consolas", 11), width=140)
        self.roi_menu.pack(side="left", padx=4)

        ctk.CTkButton(bar, text="X DEL ROI", fg_color=C_RED, hover_color="#991b1b",
                      font=("Consolas", 11, "bold"), width=100, height=34,
                      command=self._delete_roi).pack(side="left", padx=4)

        # ── ROI size inputs (cm) ─────────────────────────────────
        ctk.CTkFrame(bar, width=1, fg_color=C_BORDER).pack(side="left", fill="y", padx=8, pady=8)
        ctk.CTkLabel(bar, text="W:", font=("Consolas", 10, "bold"),
                     text_color=C_SUBTEXT).pack(side="left", padx=(2, 0))
        self.roi_w_entry = ctk.CTkEntry(bar, width=54, font=("Consolas", 11),
                                        placeholder_text="auto", justify="center")
        self.roi_w_entry.pack(side="left", padx=(2, 4))
        ctk.CTkLabel(bar, text="H:", font=("Consolas", 10, "bold"),
                     text_color=C_SUBTEXT).pack(side="left", padx=(2, 0))
        self.roi_h_entry = ctk.CTkEntry(bar, width=54, font=("Consolas", 11),
                                        placeholder_text="auto", justify="center")
        self.roi_h_entry.pack(side="left", padx=(2, 4))
        ctk.CTkLabel(bar, text="cm", font=("Consolas", 9),
                     text_color=C_SUBTEXT).pack(side="left", padx=(0, 4))

        # ── px/cm calibration ────────────────────────────────────
        ctk.CTkLabel(bar, text="px/cm:", font=("Consolas", 9),
                     text_color=C_SUBTEXT).pack(side="left", padx=(4, 0))
        self.px_per_cm_entry = ctk.CTkEntry(bar, width=44, font=("Consolas", 11),
                                             justify="center")
        self.px_per_cm_entry.pack(side="left", padx=(2, 4))
        self.px_per_cm_entry.insert(0, str(DEFAULT_PX_PER_CM))

        self.roi_mode_lbl = ctk.CTkLabel(bar, text="",
                                         font=("Consolas", 11, "bold"),
                                         text_color=C_ORANGE)
        self.roi_mode_lbl.pack(side="left", padx=10)

        self.qr_gate_lbl = ctk.CTkLabel(bar, text="",
                                        font=("Consolas", 11, "bold"),
                                        text_color="#6b7280")
        self.qr_gate_lbl.pack(side="right", padx=8)

        ctk.CTkButton(bar, text="XUẤT TỔNG HỢP",
                      fg_color="#14532d", hover_color="#166534",
                      font=("Consolas", 10, "bold"), width=140, height=34,
                      command=self.export_combined_excel).pack(side="right", padx=4)

        # ── Body: [Camera] [Kho hàng] [Hàng đã lấy] ─────────────
        body = ctk.CTkFrame(root, fg_color=C_BG)
        body.pack(fill="both", expand=True, padx=8, pady=(4, 8))
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=0, minsize=280)
        body.grid_columnconfigure(2, weight=0, minsize=280)
        body.grid_rowconfigure(0, weight=1)

        # Camera (bên trái, chiếm nhiều nhất)
        cam_card = ctk.CTkFrame(body, fg_color=C_CARD, corner_radius=10)
        cam_card.grid(row=0, column=0, sticky="nsew", padx=(0, 4))
        cam_card.grid_rowconfigure(0, weight=1)
        cam_card.grid_columnconfigure(0, weight=1)

        self.cam_label = ctk.CTkLabel(cam_card, text="NO CAMERA",
                                      font=("Consolas", 18), text_color=C_SUBTEXT)
        self.cam_label.grid(row=0, column=0, sticky="nsew", pady=4, padx=4)
        self.cam_label.bind("<Button-1>",        self._roi_press)
        self.cam_label.bind("<B1-Motion>",       self._roi_drag)
        self.cam_label.bind("<ButtonRelease-1>", self._roi_release)

        # ── Cột DANH SÁCH KHO HÀNG ───────────────────────────────
        self._inv_frame = self._make_list_panel(
            body,
            title=" KHO HÀNG",
            title_color=C_ACCENT,
            clear_cmd=self._clear_inventory,
            export_cmd=self._export_inventory_excel,
        )
        self._inv_frame.grid(row=0, column=1, sticky="nsew", padx=(0, 4))

        self._inv_count_lbl = self._inv_frame._count_lbl
        self._inv_scroll    = self._inv_frame._scroll
        self._inv_rows: dict = {}

        # ── Cột HÀNG ĐÃ LẤY ─────────────────────────────────────
        self._taken_frame = self._make_list_panel(
            body,
            title="✅  HÀNG ĐÃ LẤY",
            title_color=C_YELLOW,
            clear_cmd=self._clear_taken,
            export_cmd=self._export_taken_excel,
        )
        self._taken_frame.grid(row=0, column=2, sticky="nsew", padx=(0, 0))

        self._taken_count_lbl = self._taken_frame._count_lbl
        self._taken_scroll    = self._taken_frame._scroll
        self._taken_rows: dict = {}

    def _make_list_panel(self, parent, title: str, title_color: str, clear_cmd, export_cmd=None):
        """Tạo 1 panel cột danh sách (dùng chung cho cả 2 cột)."""
        panel = ctk.CTkFrame(parent, fg_color=C_PANEL, corner_radius=10, width=270)
        panel.grid_propagate(False)

        header = ctk.CTkFrame(panel, fg_color="#0d1f3c", corner_radius=8, height=44)
        header.pack(fill="x", padx=8, pady=(8, 4))
        header.pack_propagate(False)

        ctk.CTkLabel(header, text=title,
                     font=("Consolas", 10, "bold"),
                     text_color=title_color).pack(side="left", padx=10, pady=8)

        count_lbl = ctk.CTkLabel(header, text="0 mục",
                                  font=("Consolas", 10), text_color=C_SUBTEXT)
        count_lbl.pack(side="right", padx=8)
        panel._count_lbl = count_lbl

        # ── Hàng nút XOÁ + XUẤT EXCEL ────────────────────────────
        btn_row = ctk.CTkFrame(panel, fg_color="transparent")
        btn_row.pack(fill="x", padx=8, pady=(0, 4))

        ctk.CTkButton(btn_row, text="DEL  XOÁ",
                      fg_color="#1a1a2e", hover_color=C_RED,
                      font=("Consolas", 10, "bold"), height=28,
                      command=clear_cmd).pack(side="left", fill="x", expand=True, padx=(0, 2))

        if export_cmd:
            ctk.CTkButton(btn_row, text="EXCEL",
                          fg_color="#14532d", hover_color="#166534",
                          font=("Consolas", 10, "bold"), height=28,
                          command=export_cmd).pack(side="left", fill="x", expand=True, padx=(2, 0))

        scroll = ctk.CTkScrollableFrame(panel, fg_color="transparent", corner_radius=6)
        scroll.pack(fill="both", expand=True, padx=6, pady=(0, 8))
        panel._scroll = scroll

        return panel

    # ============================================================
    # CAMERA
    # ============================================================

    def _open_camera(self):
        idx = int(self.cam_menu.get())
        # Kiểm tra camera đang được Control Xe dùng
        car_camera_idx = self.shared.get("car_camera_index")
        if car_camera_idx is not None and int(car_camera_idx) == idx:
            car_tab = self.shared.get("car_control_tab")
            if car_tab is not None and getattr(car_tab, "running", False):
                self.cam_label.configure(text=f"CAM {idx} DANG DUOC CONTROL XE DUNG", text_color=C_RED)
                return

        # FIX reconnect: nếu camera đang chạy thì close sạch trước rồi mới open lại
        if self.running:
            self.close_camera()

        self.cam_label.configure(text=f"Dang mo camera {idx}...", text_color=C_SUBTEXT, image=None)
        self.shared["warehouse_camera_index"] = idx

        if not self.camera.open(idx):
            self.cam_label.configure(text="CAMERA OPEN FAILED", text_color=C_RED)
            self.shared.pop("warehouse_camera_index", None)
            return

        self.running = True
        self._ui_generation += 1          # vô hiệu hoá vòng after cũ nếu còn sót
        self._worker_thread = threading.Thread(target=self._frame_worker, daemon=True)
        self._worker_thread.start()
        self._schedule_ui_update(self._ui_generation)

    def close_camera(self):
        # Bước 1: báo worker thread dừng và huỷ after callback
        self.running = False
        self._ui_generation += 1      # hủy vòng after đang chờ (nếu có)
        if self._ui_after_id is not None:
            try:
                self.parent.after_cancel(self._ui_after_id)
            except Exception:
                pass
            self._ui_after_id = None
        self._last_photo = None
        self._last_tk_photo = None    # FIX pyimage v2: xóa ref PhotoImage thô
        self.shared.pop("warehouse_camera_index", None)

        # Bước 2: đợi worker thread thoát trước khi đóng camera
        # FIX: không join -> worker vẫn đang read() khi camera.close() -> V4L2 handle leak
        old_thread = self._worker_thread
        self._worker_thread = None
        if old_thread is not None and old_thread is not threading.current_thread():
            try:
                old_thread.join(timeout=0.8)
            except Exception:
                pass

        # Bước 3: đóng camera và tạo CameraManager mới
        # FIX: không tạo mới -> open() lần sau dùng object cũ đã dirty -> fail
        try:
            self.camera.close()
        except Exception:
            pass
        try:
            import time as _time
            _time.sleep(0.12)   # cho V4L2 release device handle trước khi open lại
        except Exception:
            pass
        self.camera = CameraManager()

        # Bước 4: xoá queue và cập nhật UI
        try:
            while True:
                self._frame_queue.get_nowait()
        except queue.Empty:
            pass
        # FIX pyimage: xóa image reference TRƯỚC khi configure để tránh
        # "pyimage N doesn't exist" khi CTkImage đã bị GC
        try:
            if self.cam_label.winfo_exists():
                self.cam_label.image = None
                self.cam_label.configure(image=None, text="CAMERA CLOSED", text_color=C_SUBTEXT)
        except Exception:
            pass

    # ============================================================
    # WORKER THREAD
    # ============================================================

    def _frame_worker(self):
        try:
            cv2 = get_cv2()
        except Exception as exc:
            self.running = False
            try:
                self.cam_label.configure(text=f"OPEN CV ERROR: {exc}", text_color=C_RED)
            except Exception:
                pass
            return
        while self.running:
            frame = self.camera.read()
            if frame is None:
                threading.Event().wait(0.03)
                continue

            display = frame.copy()
            self._last_frame_shape = frame.shape[:2]   # (h, w) – cập nhật mỗi frame
            # QR đang thấy trong frame này: qr -> roi_name
            current_visible: dict = {}

            for roi in self.roi_manager.rois:
                x1, y1, x2, y2 = roi["x1"], roi["y1"], roi["x2"], roi["y2"]
                cv2.rectangle(display, (x1, y1), (x2, y2), (0, 210, 255), 2)

                roi_label = roi["name"]
                mapping = self.shared.get("mapping")
                if mapping:
                    pick = mapping.get_pick_for_roi(roi["name"])
                    if pick:
                        roi_label = f"{roi['name']} -> {pick}"
                cv2.putText(display, roi_label, (x1 + 4, y1 - 8),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.62, (0, 210, 255), 2)

                crop = frame[y1:y2, x1:x2]
                if crop.size == 0:
                    continue

                for qr in self._read_multi_qr(crop, cv2):
                    current_visible[qr] = roi["name"]
                    cv2.putText(display,
                                f"QR:{qr}",
                                (x1, y2 + 22 + list(current_visible).index(qr) * 22),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.62, (0, 255, 120), 2)
                    self._try_auto_map_qr(qr)

            # Cập nhật inventory + last_seen
            now = time.time()
            now_str = datetime.now().strftime("%H:%M:%S")
            with self._lock:
                # Cập nhật các QR đang thấy
                for qr, roi_name in current_visible.items():
                    self._last_seen[qr] = now
                    if qr not in self._inventory:
                        self._inventory[qr] = {
                            "roi":        roi_name,
                            "time_first": now_str,
                            "time_last":  now_str,
                            "count":      1,
                        }
                        self._inventory_order.append(qr)
                    else:
                        self._inventory[qr]["count"]     += 1
                        self._inventory[qr]["time_last"]  = now_str
                        self._inventory[qr]["roi"]        = roi_name

                # Kiểm tra QR vắng mặt > 15s -> chuyển sang "đã lấy"
                to_move = []
                for qr in list(self._inventory.keys()):
                    if qr not in current_visible:
                        last = self._last_seen.get(qr, now)
                        if now - last >= QR_ABSENT_TIMEOUT:
                            to_move.append(qr)

                for qr in to_move:
                    info = self._inventory.pop(qr)
                    if qr in self._inventory_order:
                        self._inventory_order.remove(qr)
                    self._last_seen.pop(qr, None)
                    self._taken[qr] = {
                        "roi":      info["roi"],
                        "time_in":  info["time_first"],
                        "time_out": now_str,
                    }
                    if qr not in self._taken_order:
                        self._taken_order.append(qr)

            # Cập nhật shared
            if current_visible:
                first_qr  = next(iter(current_visible))
                self.shared["last_qr"]  = first_qr
                self.shared["last_roi"] = current_visible[first_qr]

            # ── Preview ROI đang vẽ (drag-to-draw) ───────────────
            if self._roi_drawing and self._roi_drag_start and self._roi_drag_end:
                fh, fw = display.shape[:2]
                sx, sy = self._roi_drag_start   # điểm nhấn đầu = góc trên-trái
                ex, ey = self._roi_drag_end
                # Nếu nhập kích thước cố định (cm) -> preview hình chữ nhật cố định
                try:
                    fw_str = self.roi_w_entry.get().strip()
                    fh_str = self.roi_h_entry.get().strip()
                    if fw_str and fh_str:
                        ppc = self._get_px_per_cm()
                        w_cm = float(fw_str)
                        h_cm = float(fh_str)
                        if w_cm > 0 and h_cm > 0:
                            # kích thước cố định: ex/ey = sx + kích thước (display px)
                            ex = sx + int(round(w_cm * ppc))
                            ey = sy + int(round(h_cm * ppc))
                except Exception:
                    pass
                # scale toạ độ từ display label -> frame gốc (dùng kích thước thực)
                try:
                    _lw = self.cam_label.winfo_width()
                    _lh = self.cam_label.winfo_height()
                    if _lw < 10 or _lh < 10:
                        _lw, _lh = _DISPLAY_W, _DISPLAY_H
                except Exception:
                    _lw, _lh = _DISPLAY_W, _DISPLAY_H
                sx_f = int(sx * fw / _lw)
                sy_f = int(sy * fh / _lh)
                ex_f = int(ex * fw / _lw)
                ey_f = int(ey * fh / _lh)
                # Luôn vẽ từ điểm nhấn (sx_f, sy_f) về phía ex_f, ey_f
                rx1 = min(sx_f, ex_f); rx2 = max(sx_f, ex_f)
                ry1 = min(sy_f, ey_f); ry2 = max(sy_f, ey_f)
                cv2.rectangle(display, (rx1, ry1), (rx2, ry2), (0, 255, 80), 2)
                cv2.putText(display, "VE ROI",
                            (rx1 + 4, ry1 - 8),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 80), 2)
                # góc kéo dài
                for pt in [(rx1, ry1), (rx2, ry1), (rx1, ry2), (rx2, ry2)]:
                    cv2.circle(display, pt, 5, (0, 255, 80), -1)

            rgb = cv2.cvtColor(display, cv2.COLOR_BGR2RGB)
            # FIX: dùng kích thước thực của label thay vì hằng số cố định
            try:
                dw = self.cam_label.winfo_width()
                dh = self.cam_label.winfo_height()
                if dw < 10 or dh < 10:
                    dw, dh = _DISPLAY_W, _DISPLAY_H
            except Exception:
                dw, dh = _DISPLAY_W, _DISPLAY_H
            img = Image.fromarray(rgb).resize((dw, dh), Image.BILINEAR)
            try:
                self._frame_queue.put_nowait((img, dw, dh))
            except queue.Full:
                pass

    # ────────────────────────────────────────────────────────────
    def _read_multi_qr(self, crop, cv2=None) -> list:
        if cv2 is None:
            cv2 = get_cv2()
        results = []
        seen = set()
        try:
            from pyzbar.pyzbar import decode as pyzbar_decode
            for code in pyzbar_decode(crop):
                data = code.data.decode("utf-8", errors="ignore").strip()
                if data and data not in seen:
                    results.append(data); seen.add(data)
        except Exception:
            pass
        if not results:
            try:
                ok, decoded_list, _, _ = cv2.QRCodeDetector().detectAndDecodeMulti(crop)
                if ok and decoded_list:
                    for data in decoded_list:
                        data = data.strip()
                        if data and data not in seen:
                            results.append(data); seen.add(data)
            except Exception:
                try:
                    data, _, _ = cv2.QRCodeDetector().detectAndDecode(crop)
                    if data and data.strip():
                        results.append(data.strip())
                except Exception:
                    pass
        return results

    # ============================================================
    # AUTO-MAP
    # ============================================================

    def _try_auto_map_qr(self, qr_code: str):
        if qr_code in self._auto_mapped_qr:
            return
        mapping = self.shared.get("mapping")
        if not mapping:
            return
        if mapping.get_place_for_qr(qr_code):
            self._auto_mapped_qr.add(qr_code); return
        m = re.search(r"(\d+)$", qr_code)
        if not m:
            return
        n = int(m.group(1))
        for name in (f"place{n}", f"place_{n}", f"Place{n}"):
            if name in self.shared.get("positions", {}).get("place", {}):
                mapping.set_qr_place(qr_code, name)
                self._auto_mapped_qr.add(qr_code)
                print(f"[AUTO-MAP] QR '{qr_code}' -> '{name}'")
                break

    # ============================================================
    # UI UPDATE LOOP
    # ============================================================

    def _schedule_ui_update(self, generation=None):
        if generation is None:
            generation = self._ui_generation
        # Nếu generation không khớp nghĩa là đã có open/close mới → dừng vòng cũ
        if not self.running or generation != self._ui_generation:
            return
        self._do_ui_update_v2()
        # FIX pyimage: kiểm tra lại SAU _do_ui_update_v2 vì close_camera() có thể
        # đã chạy trong khoảng thời gian đó và tăng _ui_generation → không schedule thêm
        if not self.running or generation != self._ui_generation:
            return
        # Lưu after_id để close_camera() có thể cancel ngay lập tức
        self._ui_after_id = self.parent.after(
            _FRAME_INTERVAL_MS,
            lambda g=generation: self._schedule_ui_update(g)
        )

    def _do_ui_update_v2(self):
        # Gate countdown
        try:
            countdown = self.shared.get("qr_gate_countdown", -1)
            if countdown < 0:
                self.qr_gate_lbl.configure(text="", text_color="#6b7280")
            else:
                c = "#ef4444" if countdown <= 3 else "#f97316" if countdown <= 6 else "#6b7280"
                self.qr_gate_lbl.configure(text=f"... Gate: {countdown}s", text_color=c)
        except Exception:
            pass

        try:
            item = self._frame_queue.get_nowait()
            img = item[0]
            # FIX: dùng kích thước từ worker (đã khớp với label thực tế)
            dw = item[1] if len(item) > 1 else img.size[0]
            dh = item[2] if len(item) > 2 else img.size[1]
            photo = ctk.CTkImage(light_image=img, dark_image=img, size=(dw, dh))
            # Giữ reference trên cả self và label để tránh lỗi "pyimage N doesn't exist"
            # khi reconnect nhiều lần (giống fix v17 ở car_control_tab)
            self._last_photo = photo
            # FIX pyimage v2: buộc CTkImage tạo PhotoImage nội bộ NGAY BÂY GIỜ
            # trên main thread (trước khi configure), rồi giữ hard reference để
            # GC không thu hồi PhotoImage trước khi Tkinter render xong.
            # CTkImage._create_scaled_photo_images() là internal method nhưng
            # workaround an toàn nhất là lấy PhotoImage qua get_tk_image().
            try:
                tk_img = photo.cget("light_image") if hasattr(photo, "cget") else None
            except Exception:
                tk_img = None
            # Cách chắc chắn nhất: ép CTkImage render bằng cách gọi configure
            # trước, rồi lấy lại PhotoImage từ label._image attribute nội bộ.
            if self.running and self.cam_label.winfo_exists():
                self.cam_label.image = photo  # giữ ref trên widget
                try:
                    # Gán configure trong try riêng để bắt TclError "pyimage N doesn't exist"
                    self.cam_label.configure(image=photo, text="")
                    # Sau khi configure thành công, lưu thêm ref tới PhotoImage thô
                    # bên trong CTkImage để chắc chắn không bị GC giữa các frames
                    if hasattr(photo, "_light_image"):
                        self._last_tk_photo = photo._light_image
                    elif hasattr(photo, "_photo_image"):
                        self._last_tk_photo = photo._photo_image
                except Exception:
                    pass
        except queue.Empty:
            pass

        self._refresh_inventory_ui()
        self._refresh_taken_ui()

    # ============================================================
    # RENDER – KHO HÀNG
    # ============================================================

    def _refresh_inventory_ui(self):
        with self._lock:
            order = list(self._inventory_order)
            inv   = dict(self._inventory)

        for qr in list(self._inv_rows):
            if qr not in inv:
                try: self._inv_rows[qr].destroy()
                except: pass
                del self._inv_rows[qr]

        for idx, qr in enumerate(order):
            info = inv.get(qr)
            if not info:
                continue
            if qr not in self._inv_rows:
                self._inv_rows[qr] = self._make_inv_row(qr, info, idx)
            else:
                self._update_inv_row_data(self._inv_rows[qr], info)

        self._inv_count_lbl.configure(text=f"{len(order)} mục")

    def _make_inv_row(self, qr: str, info: dict, idx: int):
        bg = "#0d1f3c" if idx % 2 == 0 else "#0a1830"
        row = ctk.CTkFrame(self._inv_scroll, fg_color=bg, corner_radius=6, height=80)
        row.pack(fill="x", pady=2, padx=2)
        row.pack_propagate(False)

        ctk.CTkLabel(row, text=f"{idx+1:02d}",
                     font=("Consolas", 10, "bold"), text_color=C_SUBTEXT,
                     width=24).pack(side="left", padx=(6, 2), pady=6)

        col = ctk.CTkFrame(row, fg_color="transparent")
        col.pack(side="left", fill="both", expand=True, padx=4)

        qr_lbl = ctk.CTkLabel(col, text=f" {qr}",
                               font=("Consolas", 11, "bold"),
                               text_color=C_GREEN, anchor="w")
        qr_lbl.pack(fill="x", pady=(5, 0))

        detail = ctk.CTkLabel(col,
                               text=f"[ROI] {info['roi']}",
                               font=("Consolas", 9), text_color=C_SUBTEXT, anchor="w")
        detail.pack(fill="x")

        time_lbl = ctk.CTkLabel(col,
                                 text=f"TIME vào: {info['time_first']}",
                                 font=("Consolas", 9), text_color="#4b5563", anchor="w")
        time_lbl.pack(fill="x")

        count_lbl = ctk.CTkLabel(row, text=f"×{info['count']}",
                                  font=("Consolas", 10, "bold"),
                                  text_color=C_YELLOW, width=34)
        count_lbl.pack(side="right", padx=4)

        def make_del(q=qr):
            def _del():
                with self._lock:
                    self._inventory.pop(q, None)
                    self._last_seen.pop(q, None)
                    if q in self._inventory_order:
                        self._inventory_order.remove(q)
                if q in self._inv_rows:
                    try: self._inv_rows[q].destroy()
                    except: pass
                    del self._inv_rows[q]
                self._inv_count_lbl.configure(text=f"{len(self._inventory_order)} mục")
            return _del

        ctk.CTkButton(row, text="X", width=24, height=24,
                      fg_color="#1a1a2e", hover_color=C_RED,
                      font=("Consolas", 10), text_color="#ff6b6b",
                      command=make_del(qr)).pack(side="right", padx=(0, 4))

        row._detail    = detail
        row._time_lbl  = time_lbl
        row._count_lbl = count_lbl
        return row

    def _update_inv_row_data(self, row, info: dict):
        try:
            row._detail.configure(text=f"[ROI] {info['roi']}")
            row._time_lbl.configure(text=f"TIME vào: {info['time_first']}")
            row._count_lbl.configure(text=f"×{info['count']}")
        except Exception:
            pass

    def _clear_inventory(self):
        with self._lock:
            self._inventory.clear()
            self._inventory_order.clear()
            self._last_seen.clear()
        for w in list(self._inv_rows.values()):
            try: w.destroy()
            except: pass
        self._inv_rows.clear()
        self._inv_count_lbl.configure(text="0 mục")

    # ============================================================
    # RENDER – HÀNG ĐÃ LẤY
    # ============================================================

    def _refresh_taken_ui(self):
        with self._lock:
            order = list(self._taken_order)
            taken = dict(self._taken)

        # Xoá widget cho QR đã bị remove khỏi taken
        for qr in list(self._taken_rows):
            if qr not in taken:
                try: self._taken_rows[qr].destroy()
                except: pass
                del self._taken_rows[qr]

        for idx, qr in enumerate(order):
            info = taken.get(qr)
            if not info:
                continue
            if qr not in self._taken_rows:
                self._taken_rows[qr] = self._make_taken_row(qr, info, idx)
            # taken rows không cần update (dữ liệu tĩnh)

        self._taken_count_lbl.configure(text=f"{len(order)} mục")

    def _make_taken_row(self, qr: str, info: dict, idx: int):
        bg = "#1a1500" if idx % 2 == 0 else "#120f00"
        row = ctk.CTkFrame(self._taken_scroll, fg_color=bg, corner_radius=6, height=80)
        row.pack(fill="x", pady=2, padx=2)
        row.pack_propagate(False)

        ctk.CTkLabel(row, text=f"{idx+1:02d}",
                     font=("Consolas", 10, "bold"), text_color=C_SUBTEXT,
                     width=24).pack(side="left", padx=(6, 2), pady=6)

        col = ctk.CTkFrame(row, fg_color="transparent")
        col.pack(side="left", fill="both", expand=True, padx=4)

        ctk.CTkLabel(col, text=f"✅  {qr}",
                     font=("Consolas", 11, "bold"),
                     text_color=C_YELLOW, anchor="w").pack(fill="x", pady=(5, 0))

        ctk.CTkLabel(col, text=f"[ROI] {info['roi']}",
                     font=("Consolas", 9), text_color=C_SUBTEXT, anchor="w").pack(fill="x")

        ctk.CTkLabel(col,
                     text=f"TIME {info['time_in']} -> {info['time_out']}",
                     font=("Consolas", 9), text_color="#6b4b00", anchor="w").pack(fill="x")

        def make_del(q=qr):
            def _del():
                with self._lock:
                    self._taken.pop(q, None)
                    if q in self._taken_order:
                        self._taken_order.remove(q)
                if q in self._taken_rows:
                    try: self._taken_rows[q].destroy()
                    except: pass
                    del self._taken_rows[q]
                self._taken_count_lbl.configure(text=f"{len(self._taken_order)} mục")
            return _del

        ctk.CTkButton(row, text="X", width=24, height=24,
                      fg_color="#1a1a2e", hover_color=C_RED,
                      font=("Consolas", 10), text_color="#ff6b6b",
                      command=make_del(qr)).pack(side="right", padx=(0, 4))

        return row

    # ============================================================
    # XUẤT EXCEL
    # ============================================================

    def _get_excel_save_path(self, default_name: str) -> str | None:
        """Mở hộp thoại chọn nơi lưu file Excel. Trả về path hoặc None nếu huỷ."""
        path = filedialog.asksaveasfilename(
            title="Lưu file Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx"), ("Tất cả file", "*.*")],
        )
        return path if path else None

    def _build_excel_styles(self, wb):
        """Trả về dict các style dùng chung."""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        thin = Side(style="thin", color="B0B8C1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return {
            "title_font":    Font(name="Arial", bold=True, size=14, color="FFFFFF"),
            "title_fill_inv": PatternFill("solid", start_color="1A5276"),
            "title_fill_tak": PatternFill("solid", start_color="7D6608"),
            "header_font":   Font(name="Arial", bold=True, size=10, color="FFFFFF"),
            "header_fill_inv": PatternFill("solid", start_color="2E86C1"),
            "header_fill_tak": PatternFill("solid", start_color="D4AC0D"),
            "even_fill":     PatternFill("solid", start_color="EBF5FB"),
            "even_fill_tak": PatternFill("solid", start_color="FEF9E7"),
            "odd_fill":      PatternFill("solid", start_color="FDFEFE"),
            "center":        Alignment(horizontal="center", vertical="center"),
            "left":          Alignment(horizontal="left",   vertical="center"),
            "border":        border,
            "normal_font":   Font(name="Arial", size=10),
            "bold_font":     Font(name="Arial", bold=True, size=10),
        }

    def _write_inventory_sheet(self, ws, styles, timestamp: str):
        """Ghi dữ liệu Kho Hàng vào worksheet."""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        with self._lock:
            order = list(self._inventory_order)
            inv   = dict(self._inventory)

        s = styles
        # Tiêu đề lớn
        ws.merge_cells("A1:F1")
        ws["A1"] = f" DANH SÁCH KHO HÀNG  -  Xuất lúc {timestamp}"
        ws["A1"].font      = s["title_font"]
        ws["A1"].fill      = s["title_fill_inv"]
        ws["A1"].alignment = s["center"]
        ws.row_dimensions[1].height = 30

        # Header
        headers = ["STT", "Mã QR", "Vị trí ROI", "Thời gian vào", "Cập nhật lần cuối", "Số lần quét"]
        widths  = [6,      24,       18,            16,              16,                  14]
        for col, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = s["header_font"]
            cell.fill      = s["header_fill_inv"]
            cell.alignment = s["center"]
            cell.border    = s["border"]
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 22

        # Dữ liệu
        for idx, qr in enumerate(order):
            info = inv.get(qr, {})
            row_num = idx + 3
            fills = [s["even_fill"] if idx % 2 == 0 else s["odd_fill"]] * 6
            values = [
                idx + 1,
                qr,
                info.get("roi", ""),
                info.get("time_first", ""),
                info.get("time_last", ""),
                info.get("count", 0),
            ]
            aligns = [s["center"], s["left"], s["center"], s["center"], s["center"], s["center"]]
            for col, (val, fill, align) in enumerate(zip(values, fills, aligns), start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font      = s["normal_font"]
                cell.fill      = fill
                cell.alignment = align
                cell.border    = s["border"]
            ws.row_dimensions[row_num].height = 20

        # Tổng
        total_row = len(order) + 3
        ws.merge_cells(f"A{total_row}:E{total_row}")
        ws.cell(row=total_row, column=1, value="TỔNG SỐ MẶT HÀNG TRONG KHO").font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=total_row, column=1).alignment = s["center"]
        ws.cell(row=total_row, column=6, value=len(order)).font = Font(name="Arial", bold=True, size=10, color="1A5276")
        ws.cell(row=total_row, column=6).alignment = s["center"]

        # Freeze panes
        ws.freeze_panes = "A3"
        return len(order)

    def _write_taken_sheet(self, ws, styles, timestamp: str):
        """Ghi dữ liệu Hàng Đã Lấy vào worksheet."""
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        with self._lock:
            order = list(self._taken_order)
            taken = dict(self._taken)

        s = styles
        ws.merge_cells("A1:F1")
        ws["A1"] = f"✅  HÀNG ĐÃ LẤY  -  Xuất lúc {timestamp}"
        ws["A1"].font      = s["title_font"]
        ws["A1"].fill      = s["title_fill_tak"]
        ws["A1"].alignment = s["center"]
        ws.row_dimensions[1].height = 30

        headers = ["STT", "Mã QR", "Vị trí ROI", "Thời gian vào kho", "Thời gian lấy ra", "Thời gian lưu kho"]
        widths  = [6,      24,       18,            18,                  18,                  18]
        for col, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font      = s["header_font"]
            cell.fill      = s["header_fill_tak"]
            cell.alignment = s["center"]
            cell.border    = s["border"]
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 22

        for idx, qr in enumerate(order):
            info = taken.get(qr, {})
            row_num = idx + 3
            fill = s["even_fill_tak"] if idx % 2 == 0 else s["odd_fill"]
            # Tính thời gian lưu kho (dạng chuỗi HH:MM:SS)
            time_in  = info.get("time_in",  "")
            time_out = info.get("time_out", "")
            duration = ""
            try:
                fmt = "%H:%M:%S"
                dt_in  = datetime.strptime(time_in,  fmt)
                dt_out = datetime.strptime(time_out, fmt)
                delta  = dt_out - dt_in
                if delta.total_seconds() < 0:
                    delta = -delta
                h, rem = divmod(int(delta.total_seconds()), 3600)
                m, sec = divmod(rem, 60)
                duration = f"{h:02d}:{m:02d}:{sec:02d}"
            except Exception:
                duration = ""

            values = [idx + 1, qr, info.get("roi", ""), time_in, time_out, duration]
            aligns = [s["center"], s["left"], s["center"], s["center"], s["center"], s["center"]]
            for col, (val, align) in enumerate(zip(values, aligns), start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font      = s["normal_font"]
                cell.fill      = fill
                cell.alignment = align
                cell.border    = s["border"]
            ws.row_dimensions[row_num].height = 20

        total_row = len(order) + 3
        ws.merge_cells(f"A{total_row}:E{total_row}")
        ws.cell(row=total_row, column=1, value="TỔNG SỐ HÀNG ĐÃ LẤY").font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=total_row, column=1).alignment = s["center"]
        ws.cell(row=total_row, column=6, value=len(order)).font = Font(name="Arial", bold=True, size=10, color="7D6608")
        ws.cell(row=total_row, column=6).alignment = s["center"]

        ws.freeze_panes = "A3"
        return len(order)

    def _export_inventory_excel(self):
        """Xuất danh sách Kho Hàng ra file Excel."""
        try:
            from openpyxl import Workbook
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = self._get_excel_save_path(f"KhoHang_{timestamp}.xlsx")
            if not path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Kho Hàng"
            styles = self._build_excel_styles(wb)
            count = self._write_inventory_sheet(ws, styles, datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
            wb.save(path)
            messagebox.showinfo("Xuất Excel thành công",
                                f"✅ Đã xuất {count} mục KHO HÀNG\nFILE {path}")
        except Exception as e:
            messagebox.showerror("Lỗi xuất Excel", f"Không thể xuất file:\n{e}")

    def _export_taken_excel(self):
        """Xuất danh sách Hàng Đã Lấy ra file Excel."""
        try:
            from openpyxl import Workbook
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = self._get_excel_save_path(f"HangDaLay_{timestamp}.xlsx")
            if not path:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "Hàng Đã Lấy"
            styles = self._build_excel_styles(wb)
            count = self._write_taken_sheet(ws, styles, datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
            wb.save(path)
            messagebox.showinfo("Xuất Excel thành công",
                                f"✅ Đã xuất {count} mục HÀNG ĐÃ LẤY\nFILE {path}")
        except Exception as e:
            messagebox.showerror("Lỗi xuất Excel", f"Không thể xuất file:\n{e}")

    def export_combined_excel(self):
        """Xuất tổng hợp cả Kho Hàng + Hàng Đã Lấy vào 1 file Excel 2 sheet."""
        try:
            from openpyxl import Workbook
            timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = self._get_excel_save_path(f"ThongKeKho_{timestamp_str}.xlsx")
            if not path:
                return
            wb = Workbook()
            now_label = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            styles = self._build_excel_styles(wb)

            ws_inv = wb.active
            ws_inv.title = "Kho Hàng"
            cnt_inv = self._write_inventory_sheet(ws_inv, styles, now_label)

            ws_tak = wb.create_sheet("Hàng Đã Lấy")
            cnt_tak = self._write_taken_sheet(ws_tak, styles, now_label)

            wb.save(path)
            messagebox.showinfo(
                "Xuất Excel tổng hợp thành công",
                f"✅ Kho Hàng: {cnt_inv} mục\n"
                f"✅ Hàng Đã Lấy: {cnt_tak} mục\n"
                f"FILE {path}"
            )
        except Exception as e:
            messagebox.showerror("Lỗi xuất Excel", f"Không thể xuất file:\n{e}")

    def _clear_taken(self):
        with self._lock:
            self._taken.clear()
            self._taken_order.clear()
        for w in list(self._taken_rows.values()):
            try: w.destroy()
            except: pass
        self._taken_rows.clear()
        self._taken_count_lbl.configure(text="0 mục")

    # ============================================================
    # ROI – drag-to-draw
    # ============================================================

    def _enable_roi(self):
        """Kích hoạt chế độ vẽ ROI: nhấn giữ + kéo chuột để vẽ khung."""
        self.adding_roi = True
        self._roi_drag_start = None
        self._roi_drag_end   = None
        self._roi_drawing    = False
        self.roi_mode_lbl.configure(
            text="EDIT NHẤN GIỮ & KÉO để vẽ khung ROI  |  ESC để huỷ",
            text_color=C_ORANGE
        )
        # Cho phép ESC huỷ chế độ vẽ
        self.parent.winfo_toplevel().bind("<Escape>", self._cancel_roi, add="+")

    def _cancel_roi(self, _event=None):
        self.adding_roi   = False
        self._roi_drawing = False
        self._roi_drag_start = None
        self._roi_drag_end   = None
        self.roi_mode_lbl.configure(text="X Đã huỷ vẽ ROI", text_color=C_SUBTEXT)
        self.parent.after(1500, lambda: self.roi_mode_lbl.configure(text=""))

    # ── Mouse events ─────────────────────────────────────────────

    def _roi_press(self, event):
        if not self.adding_roi:
            return
        self._roi_drag_start = (event.x, event.y)
        self._roi_drag_end   = (event.x, event.y)
        self._roi_drawing    = True

    def _roi_drag(self, event):
        if not self._roi_drawing:
            return
        self._roi_drag_end = (event.x, event.y)

    def _get_px_per_cm(self) -> float:
        """Trả về số pixel trên 1 cm (từ ô px/cm, mặc định DEFAULT_PX_PER_CM)."""
        try:
            v = float(self.px_per_cm_entry.get().strip())
            return v if v > 0 else DEFAULT_PX_PER_CM
        except Exception:
            return DEFAULT_PX_PER_CM

    def _get_roi_fixed_size(self):
        """Trả về (w_px, h_px) nếu người dùng nhập W/H theo cm, ngược lại trả (None, None)."""
        try:
            w_str = self.roi_w_entry.get().strip()
            h_str = self.roi_h_entry.get().strip()
            if not w_str or not h_str:
                return None, None
            w_cm = float(w_str)
            h_cm = float(h_str)
            if w_cm <= 0 or h_cm <= 0:
                return None, None
            ppc = self._get_px_per_cm()
            w_px = int(round(w_cm * ppc))
            h_px = int(round(h_cm * ppc))
            return w_px, h_px
        except Exception:
            return None, None

    def _roi_release(self, event):
        if not self._roi_drawing or self._roi_drag_start is None:
            return
        self._roi_drawing = False

        x1, y1 = self._roi_drag_start   # điểm nhấn đầu tiên = góc trên-trái khung ROI
        x2, y2 = event.x, event.y

        # Nếu có nhập kích thước cố định (cm) -> dùng điểm nhấn làm góc trên-trái,
        # áp dụng kích thước W×H đã nhập (đã chuyển cm -> px display)
        fixed_w, fixed_h = self._get_roi_fixed_size()
        if fixed_w and fixed_h:
            rx1 = x1
            ry1 = y1
            rx2 = rx1 + fixed_w
            ry2 = ry1 + fixed_h
            # Clamp vào trong display (dùng kích thước thực của label)
            try:
                lw_clamp = self.cam_label.winfo_width()
                lh_clamp = self.cam_label.winfo_height()
                if lw_clamp < 10: lw_clamp = _DISPLAY_W
                if lh_clamp < 10: lh_clamp = _DISPLAY_H
            except Exception:
                lw_clamp, lh_clamp = _DISPLAY_W, _DISPLAY_H
            rx2 = min(rx2, lw_clamp)
            ry2 = min(ry2, lh_clamp)
        else:
            # Kéo tự do: góc trên-trái = điểm nhấn đầu, không đảo min/max
            rx1, ry1 = x1, y1
            rx2, ry2 = x2, y2
            # Đảm bảo rx2 > rx1 và ry2 > ry1
            if rx2 < rx1: rx1, rx2 = rx2, rx1
            if ry2 < ry1: ry1, ry2 = ry2, ry1

        # Bỏ qua nếu khung quá nhỏ (dưới 15×15 px display)
        if (rx2 - rx1) < 15 or (ry2 - ry1) < 15:
            self.roi_mode_lbl.configure(text="! Khung quá nhỏ, vẽ lại!", text_color=C_RED)
            self.parent.after(1500, lambda: self.roi_mode_lbl.configure(
                text="EDIT NHẤN GIỮ & KÉO để vẽ khung ROI  |  ESC để huỷ",
                text_color=C_ORANGE))
            self._roi_drag_start = None
            self._roi_drag_end   = None
            return

        idx  = len(self.roi_manager.rois) + 1
        name = f"ROI_{idx}"

        # ── Scale toạ độ display label -> frame gốc trước khi lưu ─
        fh, fw = self._last_frame_shape
        try:
            lw = self.cam_label.winfo_width()
            lh = self.cam_label.winfo_height()
            if lw < 10 or lh < 10:
                lw, lh = _DISPLAY_W, _DISPLAY_H
        except Exception:
            lw, lh = _DISPLAY_W, _DISPLAY_H
        sx1 = int(rx1 * fw / lw)
        sy1 = int(ry1 * fh / lh)
        sx2 = int(rx2 * fw / lw)
        sy2 = int(ry2 * fh / lh)

        self.roi_manager.add_roi(name, sx1, sy1, sx2, sy2)
        self.adding_roi      = False
        self._roi_drag_start = None
        self._roi_drag_end   = None

        ppc = self._get_px_per_cm()
        w_cm = (rx2 - rx1) / ppc
        h_cm = (ry2 - ry1) / ppc
        self.roi_mode_lbl.configure(
            text=f"OK {name} đã thêm  ({w_cm:.1f}×{h_cm:.1f} cm)",
            text_color=C_GREEN
        )
        self.parent.after(2000, lambda: self.roi_mode_lbl.configure(text=""))
        self._refresh_roi_list()
        self._try_auto_map_roi(name, idx)
        try:
            self.shared["auto_tab"].refresh_data()
        except Exception:
            pass

    # ── Helpers ──────────────────────────────────────────────────

    def _try_auto_map_roi(self, roi_name: str, idx: int):
        mapping = self.shared.get("mapping")
        if not mapping or mapping.get_pick_for_roi(roi_name):
            return
        pick_dict = self.shared.get("positions", {}).get("pick", {})
        for name in (f"pick{idx}", f"pick_{idx}", f"Pick{idx}"):
            if name in pick_dict:
                mapping.set_roi_pick(roi_name, name)
                print(f"[AUTO-MAP] ROI '{roi_name}' -> '{name}'")
                break

    def _refresh_roi_list(self):
        names = [r["name"] for r in self.roi_manager.rois] or ["NO ROI"]
        self.roi_menu.configure(values=names)
        self.roi_menu.set(names[0])

    def _delete_roi(self):
        name = self.roi_menu.get()
        if name == "NO ROI":
            return
        self.roi_manager.delete_roi(name)
        mapping = self.shared.get("mapping")
        if mapping:
            mapping.remove_roi_pick(name)
        self._refresh_roi_list()
        try:
            self.shared["auto_tab"].refresh_data()
        except Exception:
            pass
